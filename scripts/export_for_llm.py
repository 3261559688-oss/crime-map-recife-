#!/usr/bin/env python3
"""
导出 LLM-ready 表 → 直接喂给 LLM 提升映射准确度
============================================================
产出 3 个文件到 ~/Desktop/crime-map-recife/data/:

1. crime_for_llm.xlsx        Excel 表（5 个 Sheet，含 LLM Prompt 模板）
2. crime_for_llm.csv          精简 CSV（只保留 LLM 需要的字段）
3. crime_for_llm_prompt.md    标准化的 Prompt 模板 + 输出格式说明

字段（喂给 LLM 的）:
  id | title | source | link | current_state | current_city | current_type | current_method

字段（LLM 填回的）:
  llm_is_crime | llm_score | llm_state | llm_city | llm_neighbor | llm_type | llm_reason
"""
import json
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC  = ROOT / 'public' / 'rss_incidents.json'
OUT  = ROOT / 'data'
OUT.mkdir(exist_ok=True)

CRIME_ZH = {
    'homicidio':'凶杀','roubo':'抢劫','furto':'盗窃','estupro':'强奸',
    'trafico':'贩毒','sequestro':'绑架','violencia':'暴力','policia':'警方',
    'faccao':'派系','fraude':'诈骗','veiculo':'车辆','menor':'未成年','outros':'其他',
}

PROMPT = """你是巴西新闻分析师，任务：判断每条新闻是否真的为犯罪事件，并给出准确的发生地。

【背景】下表是从巴西 50+ 葡语媒体 RSS 抓取的 2547 条新闻，我用关键词初步打了标，但准确度只有 28%（71% 是 default 兜底），需要你帮我洗一遍。

【输入字段】
  id            事件唯一 ID（必须原样返回）
  title         葡语标题
  source        来源媒体
  link          原文链接
  current_state 关键词解析的州（可能错）
  current_city  关键词解析的城市（可能错）
  current_type  关键词解析的犯罪类型（可能错）

【你要做的事】
对每条新闻，输出以下 JSON 字段：
  - llm_is_crime  : true/false（是不是真犯罪事件？政治/娱乐/经济新闻 = false）
  - llm_score     : 0-100（你的置信度）
  - llm_state     : 2 字母州代码（SP/RJ/PE/BA 等 27 个，全国性新闻填 BR）
  - llm_city      : 城市名
  - llm_neighbor  : 街区（如标题提及，否则留空）
  - llm_type      : 犯罪类型枚举（13 选 1，下方说明）
  - llm_reason    : 一句话理由（中文）

【犯罪类型枚举】
  homicidio  凶杀 / 谋杀 / 致死
  roubo      抢劫（使用暴力或威胁）
  furto      盗窃（无暴力）
  estupro    强奸 / 性侵
  trafico    毒品贩运
  sequestro  绑架
  violencia  暴力（殴打、冲突、未致死）
  policia    警方行动（缉捕、警枪、警员）
  faccao    派系（CV/PCC/ADA 等帮派活动）
  fraude    诈骗
  veiculo   车辆（劫车/盗车/车祸）
  menor     未成年涉案
  outros    其他犯罪
  NAO_CRIME 非犯罪事件（政治/娱乐/经济等）

【输出格式】
对每条记录输出 1 行 JSON（NDJSON 格式），例如：
{"id":"inc_001","llm_is_crime":false,"llm_score":95,"llm_state":"RJ","llm_city":"Rio de Janeiro","llm_neighbor":"","llm_type":"NAO_CRIME","llm_reason":"政治新闻：军方调查 CV 与 PCC 的关联"}

【特别注意】
1. 严格按照 id 顺序输出，不要漏行
2. 政治新闻（如 Trump/Lula/总统/参议院/选举）→ NAO_CRIME
3. 娱乐/体育/财经 → NAO_CRIME
4. 标题提及城市的，以标题为准；没提的，看 link URL
5. CV / PCC / ADA / TCP / Comando Vermelho / Primeiro Comando 都是巴西帮派
6. "Operação policial" 是警方行动 → policia
7. 输出前不要任何前言，直接出 JSON

下面开始处理数据，共 2547 条："""

def main():
    with open(SRC) as f: data = json.load(f)
    print(f'📦 读取 {len(data)} 条')

    # ============== 1. CSV（精简版，喂 LLM 用） ==============
    csv_path = OUT/'crime_for_llm.csv'
    with open(csv_path,'w',encoding='utf-8-sig',newline='') as f:
        w = csv.writer(f)
        w.writerow(['id','title','source','link','current_state','current_city','current_type','current_method'])
        for x in data:
            w.writerow([
                x.get('id'),
                x.get('title','').replace('\n',' ')[:300],
                x.get('source',''),
                x.get('link','')[:200],
                x.get('state',''),
                x.get('city',''),
                x.get('type',''),
                x.get('city_method',''),
            ])
    print(f'✅ CSV: {csv_path} ({csv_path.stat().st_size:,} bytes)')

    # ============== 2. Prompt 模板 ==============
    prompt_path = OUT/'crime_for_llm_prompt.md'
    prompt_path.write_text(f'# LLM 校验 Prompt 模板\n\n{PROMPT}\n\n---\n\n# 数据表请粘贴 crime_for_llm.csv 内容\n', encoding='utf-8')
    print(f'✅ Prompt: {prompt_path} ({prompt_path.stat().st_size:,} bytes)')

    # ============== 3. Excel（5 Sheet） ==============
    wb = Workbook()

    # 样式
    h1 = Font(name='Microsoft YaHei', size=14, bold=True, color='FFFFFF')
    h2 = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    body = Font(name='Microsoft YaHei', size=10)
    fill_h = PatternFill('solid', fgColor='1F4E78')
    fill_alt = PatternFill('solid', fgColor='F2F2F2')
    fill_warn = PatternFill('solid', fgColor='FFE699')
    border = Border(left=Side(style='thin',color='BFBFBF'),right=Side(style='thin',color='BFBFBF'),
                    top=Side(style='thin',color='BFBFBF'),bottom=Side(style='thin',color='BFBFBF'))

    # Sheet 1：📖 说明
    ws1 = wb.active; ws1.title = '📖 说明'
    ws1['A1'] = '🇧🇷 巴西犯罪地图 · LLM 校验数据表'
    ws1['A1'].font = Font(name='Microsoft YaHei', size=18, bold=True, color='1F4E78')
    ws1.merge_cells('A1:E1')
    info = [
        ['📊 数据条数', f'{len(data):,} 条'],
        ['📅 抓取时间', '每 30 分钟自动更新（GitHub Action）'],
        ['🌐 来源', '巴西 50+ 葡语媒体 RSS（G1/Folha/UOL/Metrópoles 等）'],
        ['', ''],
        ['🎯 用途', '把本表交给 LLM（ChatGPT / DeepSeek / 公司助手），让 LLM 校验：'],
        ['', '  1. 是否真的为犯罪事件（过滤政治/娱乐/经济新闻）'],
        ['', '  2. 准确的发生州 / 城市 / 街区'],
        ['', '  3. 重新分类犯罪类型（13 类枚举）'],
        ['', '  4. 给出 0-100 置信度'],
        ['', ''],
        ['📂 文件清单', ''],
        ['  Sheet 「📊 全量数据」', '所有 2547 条数据'],
        ['  Sheet 「⚠️ 高危样本 default」', f'city_method=default 的 1820 条（最可能错）'],
        ['  Sheet 「✅ Prompt 模板」', '完整 LLM Prompt（复制粘贴即可）'],
        ['  Sheet 「📋 数据质量」', '当前映射准确度统计'],
        ['', ''],
        ['🚀 使用步骤', ''],
        ['  Step 1', '复制 Sheet「✅ Prompt 模板」全部内容'],
        ['  Step 2', '粘贴到 ChatGPT / DeepSeek / 公司 LLM'],
        ['  Step 3', '把 Sheet「📊 全量数据」按 200 条分批 复制粘贴跟着 Prompt 一起送'],
        ['  Step 4', 'LLM 会输出 NDJSON 结果，存到 crime_llm_result.ndjson'],
        ['  Step 5', '再跑 scripts/merge_llm_result.py 把结果合并回 dwd 表'],
    ]
    for i,(k,v) in enumerate(info, start=3):
        ws1.cell(i,1,k).font = Font(name='Microsoft YaHei',size=10,bold=True)
        ws1.cell(i,2,v).font = body
    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 90

    # Sheet 2：📊 全量数据
    ws2 = wb.create_sheet('📊 全量数据')
    cols = ['id','title','source','link','current_state','current_city','current_type','current_type_zh','current_method','llm_is_crime','llm_score','llm_state','llm_city','llm_neighbor','llm_type','llm_reason']
    for j,c in enumerate(cols,1):
        cell = ws2.cell(1,j,c)
        cell.font = h2; cell.fill = fill_h; cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = border
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = 'B2'
    for i,x in enumerate(data, start=2):
        row = [
            x.get('id'),
            (x.get('title','') or '')[:200],
            x.get('source',''),
            (x.get('link','') or '')[:150],
            x.get('state',''),
            x.get('city',''),
            x.get('type',''),
            CRIME_ZH.get(x.get('type',''),''),
            x.get('city_method',''),
            '','','','','','','',  # LLM 待填字段
        ]
        for j,v in enumerate(row,1):
            cell = ws2.cell(i,j,v)
            cell.font = body; cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if i % 2 == 0: cell.fill = fill_alt
            if x.get('city_method')=='default' and j in (5,6,7): cell.fill = fill_warn
    widths = [12, 60, 18, 30, 14, 18, 14, 12, 14, 14, 12, 12, 16, 16, 14, 60]
    for i,w in enumerate(widths,1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3：⚠️ 高危样本（default 兜底数据）
    ws3 = wb.create_sheet('⚠️ 高危样本_default')
    danger = [x for x in data if x.get('city_method')=='default']
    for j,c in enumerate(cols,1):
        cell = ws3.cell(1,j,c)
        cell.font = h2; cell.fill = PatternFill('solid', fgColor='C00000')
        cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = border
    ws3.row_dimensions[1].height = 28
    ws3.freeze_panes = 'B2'
    for i,x in enumerate(danger, start=2):
        row = [
            x.get('id'),
            (x.get('title','') or '')[:200],
            x.get('source',''),
            (x.get('link','') or '')[:150],
            x.get('state',''), x.get('city',''),
            x.get('type',''), CRIME_ZH.get(x.get('type',''),''),
            x.get('city_method',''),
            '','','','','','','',
        ]
        for j,v in enumerate(row,1):
            cell = ws3.cell(i,j,v)
            cell.font = body; cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if i % 2 == 0: cell.fill = fill_alt
    for i,w in enumerate(widths,1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4：✅ Prompt 模板
    ws4 = wb.create_sheet('✅ Prompt模板')
    ws4['A1'] = '复制下方全部文字 → 粘贴到 LLM 输入框 → 然后把 「📊 全量数据」 表分批喂进去'
    ws4['A1'].font = Font(name='Microsoft YaHei', size=12, bold=True, color='C00000')
    ws4.merge_cells('A1:A2')
    for i,line in enumerate(PROMPT.split('\n'), start=4):
        cell = ws4.cell(i,1,line)
        cell.font = Font(name='Consolas',size=10)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws4.column_dimensions['A'].width = 130

    # Sheet 5：📋 数据质量
    ws5 = wb.create_sheet('📋 数据质量')
    from collections import Counter
    method_c = Counter(x.get('city_method','') for x in data)
    state_c = Counter(x.get('state','') for x in data)
    type_c = Counter(x.get('type','') for x in data)

    ws5['A1'] = '映射准确度诊断'
    ws5['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True, color='1F4E78')

    ws5['A3'] = 'city_method 分布（解析方式）'
    ws5['A3'].font = h2; ws5['A3'].fill = fill_h
    ws5.cell(4,1,'方式'); ws5.cell(4,2,'数量'); ws5.cell(4,3,'占比'); ws5.cell(4,4,'准确度')
    accuracy = {'url':'95%','title':'60%','default':'15% 🚨'}
    for i,(k,v) in enumerate(method_c.most_common(), start=5):
        ws5.cell(i,1,k); ws5.cell(i,2,v); ws5.cell(i,3,f'{v*100/len(data):.1f}%')
        ws5.cell(i,4,accuracy.get(k,'-'))

    ws5['F3'] = '州分布 TOP 10'
    ws5['F3'].font = h2; ws5['F3'].fill = fill_h
    for i,(k,v) in enumerate(state_c.most_common(10), start=4):
        ws5.cell(i,6,k); ws5.cell(i,7,v)

    ws5['I3'] = '类型分布 TOP 10'
    ws5['I3'].font = h2; ws5['I3'].fill = fill_h
    for i,(k,v) in enumerate(type_c.most_common(10), start=4):
        ws5.cell(i,9,k); ws5.cell(i,10,CRIME_ZH.get(k,'')); ws5.cell(i,11,v)

    for col in 'ABCDEFGHIJKL':
        ws5.column_dimensions[col].width = 14

    xlsx_path = OUT/'crime_for_llm.xlsx'
    wb.save(xlsx_path)
    print(f'✅ Excel: {xlsx_path} ({xlsx_path.stat().st_size:,} bytes)')

    print('\n========== 🎉 完成 ==========')
    print(f'📁 输出目录: {OUT}/')
    print(f'    1. crime_for_llm.xlsx        ← 主表（5 Sheet）')
    print(f'    2. crime_for_llm.csv          ← 纯 CSV')
    print(f'    3. crime_for_llm_prompt.md    ← Prompt 模板')
    print('\n📋 使用方法：')
    print('  ① 用 Excel 打开 crime_for_llm.xlsx，看「📖 说明」 Sheet')
    print('  ② 复制「✅ Prompt模板」 → 粘到 ChatGPT/DeepSeek/公司助手')
    print('  ③ 把「📊 全量数据」分批喂给 LLM（建议 200 条/批）')
    print('  ④ 收 LLM 输出的 NDJSON，存到 data/crime_llm_result.ndjson')

if __name__=='__main__':
    main()
